import py_cui
import py_cui.keys as KEYS
import tkinter as tk
from tkinter import filedialog
import xlrd
from openpyxl import load_workbook
import requests
import json
 
mail = {

"罗晓辉":"luoxiaohui",
"毛清":"maoqing",

}


class MouseApp:

    def __init__(self, root: py_cui.PyCUI):

        # Initialize our two widgets, a button and a mouse press log
        self.root = root
        self.button_presser = self.root.add_button('Press Me!', 0, 0)
        self.mouse_press_log = self.root.add_text_block('Mouse Presses', 0, 1, column_span=2)
        self.button_presser.add_mouse_command(KEYS.LEFT_MOUSE_CLICK, self.print_left_press_with_coords)
        #self.print_left_press_with_coords()
        
    def print_left_press_with_coords(self):
        root = tk.Tk()
        root.withdraw()
        
        file_path = filedialog. askopenfilename(title='选择你的英雄', initialdir='/', filetypes=[('xlsx','*.xlsx'),('xls','*.xls')])
        a = self.breakxlx(file_path)
        out = self.getalltext(file_path,a)
        self.msg(out)

    def breakxlx(self,path):
        
        wb = load_workbook(filename=path)        
        sheets = wb.get_sheet_names() 
        sheet_first = sheets[0]
        #获取特定的worksheet
        ws = wb.get_sheet_by_name(sheet_first)
        pool = {}
        for i in range(ws.max_row-1):
            # for ii in range(ws.max_column-1):
            text = ws.cell(row=i+1, column=2)
            if text.value != None:
                if '环节接口人评价' in text.value :
                    pool['接口人'] = {'start':i+3}
                elif 'S-A级' in text.value :
                    pool['sa'] = {'start':i+3}
                    pool['接口人']['end'] = i-2
                elif 'B级' in text.value :
                    pool['b'] = {'start':i+2}
                    pool['sa']['end']= i
                elif 'CD级' in text.value :
                    pool['cd'] = {'start':i+2}
                    pool['b']['end'] = i-1
        pool['cd']['end'] = ws.max_row-1
        return pool
        

        # #获取表格所有行和列，两者都是可迭代的
        # rows = ws.rows.__len__()
        # columns = ws.columns.__len__()
        
        # #迭代所有的行
        # for row in rows[0]:
        #     box = row
        #     print(str(box.value))
        #     self.mouse_press_log.set_text(str(box.value))


        data = xlrd.open_workbook(path)
        table = data.sheets()[0]
        maxrows = table.nrows
        maxcols = table.ncols
        pool = {}
        for i in range(maxrows):
            text = table.cell(i,1)
            #print(text)
            if '环节接口人评价' in text.value :
                pool['接口人'] = {'start':i}
            elif 'S-A级' in text.value :
                pool['sa'] = {'start':i}
                pool['接口人'] = {'end':i-3}
            elif 'B级' in text.value :
                pool['b'] = {'start':i}
                pool['sa'] = {'end':i-1}
            elif 'CD级' in text.value :
                pool['cd'] = {'start':i}
                pool['b'] = {'end':i-2}
        return pool

    def getalltext(self,path,pool):
        wb = load_workbook(filename=path)        
        sheets = wb.get_sheet_names() 
        sheet_first = sheets[0]
        #获取特定的worksheet
        ws = wb.get_sheet_by_name(sheet_first)
        res = {}
        def d(i):
            name =  ws.cell(row=i, column=2).value
            title =  ws.cell(row=i, column=3).value
            level =  ws.cell(row=i, column=4).value
            work =  ws.cell(row=i, column=5).value
            agast =  ws.cell(row=i, column=6).value
            risk=  ws.cell(row=i, column=7).value
            if name != None:
                if name in mail:
                    popo = mail[name] 
                else:
                    print("无法查找popo",name)
                    popo = None
                res[name]  = {'title':title,'level':level,'work':work,'agast':agast,'risk':risk,'popo':popo}
            return res



        for i in range(pool['接口人']['start'],pool['接口人']['end']+1):
            d(i)
        for i in range(pool['sa']['start'],pool['sa']['end']+1):
            d(i)
        for i in range(pool['b']['start'],pool['b']['end']+1):
            d(i)
        for i in range(pool['cd']['start'],pool['cd']['end']+1):
            d(i)

        for i in res:
            self.mouse_press_log.set_text(str(res))
        return res
            
    def getemail(self,name):
        pass


    @classmethod
    def send_to_users(cls,popos, msg, font='宋体', italic='0', underline='0', color='000000'):
        try:
            r = requests.post('http://hsqa.nie.netease.com/api/post_popo',
                              data={
                                  'uids': json.dumps(popos),
                                  'msg': msg,
                                  'font': font,
                                  'italic': italic,
                                  'underline': underline,
                                  'color': color
                              })
            r = r.json()
            if r['status'] == 1:
                return True
            return False
        except:
            return False



    def msg(self,dict):
        for key,value in dict.items():
            # make mssage
            message = str(value['title']) + str(key) +'\r\n你好！这是你本月的绩效：\r\n综合评价：'+str(value['level'])+'\r\n工作内容：'+str(value['work']) + '\r\n建议：'+str(value['agast'])+'\r\n风险：'+str(value['risk'])
            # send it
            self.send_to_users([value['popo']],message)

        message = "马哥，工具写完了，录入大家的邮箱之后就 ok了"
            # send it
        self.send_to_users(["macheng"],message)


root = py_cui.PyCUI(1, 3)
MouseApp(root)
root.start()
